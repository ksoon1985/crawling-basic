import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기 
ws = wb.create_sheet('오징어 게임')

# 3) 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

wb.save(r'C:\Users\kwons\Desktop\study\git_project\python_crawling\basic_crawling_1\Chapter05\참가자_data.xlsx')
